import csv, re, difflib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CORPUS="/Users/mitchel/Projects/untitled folder/digital-minds-corpus/master-corpus.csv"
def norm(s):
    s=(s or "").lower(); s=re.sub(r"[^a-z0-9 ]"," ",s); return re.sub(r"\s+"," ",s).strip()
def sim(a,b):
    a,b=norm(a),norm(b)
    if a and b and (a in b or b in a) and min(len(a),len(b))>=25: return 0.95
    return difflib.SequenceMatcher(None,a,b).ratio()

refs=list(csv.DictReader(open("seed-references.csv")))
# Holdings check runs against EVERY layer - an 'adjacent' work is still one we hold.
corpus=[x for x in csv.DictReader(open(CORPUS))]
# The personhood cluster we mine as a source is restricted to core+canon.
cluster=[x for x in corpus if x["layer"] in ("core","canon")
         and re.search(r"personhood|legal person|legal status|ai rights|rights of ai|juridical",(x["title"] or ""),re.I)]

# Manual adjudications for borderline title matches, checked by hand against the
# corpus by author AND title. These override the automatic threshold.
# Works added to the register that are NOT from the paper's reference list.
# Surfaced by desk research on prior surveys (see PRIOR-SURVEYS.md).
NOT_IN_PAPER={"jaynes2024","bayern2014","jaynes2021","attoe2023","nasir2023"}

ADJUDICATED={
 "butlin2023": ("No","Checked by hand: corpus holds other Butlin works (Identifying indicators of consciousness, 2025) but NOT the arXiv 2308.08708 report"),
 "ivanov2022": ("No","Checked by hand: corpus neuromorphic entries are different papers (Shiller 2025, Wang 2025)"),
 "salib2024a": ("No","Checked by hand: corpus holds only 'AI Rights for Human Flourishing'; this is a distinct paper"),
 "jaynes2024": ("No","Genre note: AI & Society Curmudgeon Corner is an editorially-reviewed opinion column, not a peer-reviewed research section. Kept in Literature (journal-hosted, DOI, referenced) rather than Evidence, unlike mass-media op-eds"),
 "salib2024b": ("Verify","Corpus holds 'AI Rights for Human Flourishing' (Salib & Goldstein, 2025) - may be this paper retitled on publication. Confirm against SSRN 5353214 before collecting"),
}

# ---- build union literature set ----
lit=[]
for r in refs:
    best,score=None,0
    for c in corpus:
        s=sim(r["title"],c["title"])
        if s>score: score,best=s,c
    inc = score>=0.85
    held = "Yes" if inc else "No"
    note = ""
    if r["ref_key"] in ADJUDICATED:
        held, note = ADJUDICATED[r["ref_key"]]
        inc = (held=="Yes")
    action = ("No action - already held" if held=="Yes"
              else "Verify against corpus record" if held=="Verify"
              else "Collect" if r["tier"]=="core" else "Hold (marginal)")
    lit.append(dict(ref_key=r["ref_key"], in_paper="No" if r["ref_key"] in NOT_IN_PAPER else "Yes", in_corpus=held,
        corpus_layer=best["layer"] if inc else "", tier=r["tier"], strand=r["strand"],
        authors=r["authors"], year=r["year"], title=r["title"], venue=r["venue"], type=r["type"],
        discipline=best["discipline"] if inc else "", themes=best["themes"] if inc else "",
        cited_by=best["cited_by"] if inc else "", url_or_id=r["url_or_id"],
        action=action, notes=(note or ("Not cited by the paper; surfaced by desk research" if r["ref_key"] in NOT_IN_PAPER else ""))))

paper_titles=[r["title"] for r in refs]
added=0
for c in cluster:
    if max((sim(c["title"],t) for t in paper_titles), default=0)>=0.85: continue
    added+=1
    lit.append(dict(ref_key=f"corp{added:03d}", in_paper="No", in_corpus="Yes",
        corpus_layer=c["layer"], tier="core", strand="corpus_personhood_cluster",
        authors=c["authors"], year=c["year"], title=c["title"], venue=c["venue"], type=c["type"],
        discipline=c["discipline"], themes=c["themes"], cited_by=c["cited_by"],
        url_or_id=c["doi"] or c["openalex_id"],
        action="Classify for tracker", notes="In corpus; not cited by the paper"))

# ---- evidence records ----
EV=[
("abdurasulov2025","Abdurasulov, A.","2025","The new AI arms race changing the war in Ukraine","BBC News","news","","","https://www.bbc.co.uk/news/articles/cly7jrez2jno"),
("alexander_simon2025","Alexander, H.; Simon, J.","2025","Ohio's AI Personhood Ban Risks Outlawing the Future","Ohio Capital Journal","op_ed","OH","HB 469","https://ohiocapitaljournal.com/2025/11/25/ohios-ai-personhood-ban-risks-outlawing-the-future/"),
("allianceai_a","Alliance for Secure AI Action","n.d.","The Alliance for Secure AI Action","Organisation site","advocacy","","","https://secureainow.org/"),
("allianceai_b","Alliance for Secure AI Action","n.d.","The Alliance for Secure AI Staff","Organisation site","advocacy","","","https://secureainow.org/staff/"),
("amipodcast2025","Am I? Podcast","2025","Lawmaker Explains Why He Wants to Outlaw AI Consciousness","Podcast","media","OH","HB 469","https://podcasts.apple.com/us/podcast/am-i-id1834212843?i=1000740865127"),
("aquinas","Aquinas, T.","ca. 1270","Summa Theologiae, Prima Pars, Question 93","-","primary_religious","","",""),
("ball2025a","Ball, D.","2025","Tweet reply on TN SB 1493 and the First Amendment","X (Twitter)","social","TN","HB 1455 / SB 1493","https://x.com/deanwball/status/2004678006561013931"),
("ball2025b","Ball, D.","2025","Tweet commenting on TN SB 1493","X (Twitter)","social","TN","HB 1455 / SB 1493","https://x.com/deanwball/status/2004637387402416246"),
("ball_bio","Ball, D.","n.d.","Dean W. Ball - Biography","Personal site","bio","","","https://www.deanball.com/"),
("banutah_complaint","Ban v. Utah complaint","2025","No. 250900869 (Utah Dist. Ct.)","Court filing","legal_doc","UT","HB 249","https://www.nonhumanrights.org/wp-content/uploads/2025/01/UT-Complaint.pdf"),
("banutah_dismissal","Ban v. Utah dismissal order","2025","Memorandum decision and order granting motion to dismiss","Court order","legal_doc","UT","HB 249",""),
("barker2026","Barker, D.","2026","Testimony on AB 959 before the Wisconsin Assembly Cttee on Science, Technology and AI","Wisconsin Legislature","testimony","WI","AB 959 / SB 932","https://docs.legis.wisconsin.gov/misc/lc/hearing_testimony_and_materials/2025/ab959/ab0959_2026_02_04.pdf"),
("berg2025","Berg, C.; Rosenblatt, J.","2025","If AI Becomes Conscious, We Need to Know","Wall Street Journal","op_ed","","","https://www.wsj.com/opinion/if-ai-becomes-conscious-we-need-to-know-83aa61d8"),
("berry2026","Berry, S.","2026","Testimony on Missouri HB 1746","Missouri House of Representatives","testimony","MO","HB 1746","https://documents.house.mo.gov/billtracking/bills261/witnesses/HB1746Testimony2-2.pdf"),
("bible","Bible","-","Genesis 1:26-27","-","primary_religious","","",""),
("blacks2001","Black's Law Dictionary","2001","2nd pocket edn (Bryan A. Garner, ed.)","West Group","reference_work","","",""),
("boone2022","Boone, R.","2022","Bill Would Bar Idaho's Lands and Animals From 'Personhood'","Associated Press","news","ID","HB 720","https://apnews.com/article/united-states-environment-animals-idaho-402c779428a74a6ec9e7994ee977005c"),
("brill2026","Brill, L.","2026","Testimony on AB 959 before the Wisconsin Assembly Committee","Wisconsin Legislature","testimony","WI","AB 959 / SB 932","https://docs.legis.wisconsin.gov/misc/lc/hearing_testimony_and_materials/2025/ab959/ab0959_2026_02_04.pdf"),
("claggett2025","Claggett, T.","2025","Sponsor testimony on HB 469, Ohio House Technology and Innovation Cttee","Ohio House of Representatives","testimony","OH","HB 469","https://ohiohouse.gov/committees/technology-and-innovation/bills/hb469"),
("maynard_nd","Cody Maynard for Oklahoma","n.d.","Endorsements","Campaign site","advocacy","OK","HB 3546","https://codymaynard.com/home/endorsements/"),
("colbert2025","Colbert, S.","2025","The A.I. Stock Bubble | ChatGPT, Grok, Go Erotic | Banning Human-Chatbot Marriage","YouTube","media","","","https://www.youtube.com/watch?v=o-u_ZkGBcNc"),
("copp2026","Copp, T. et al.","2026","Anthropic's AI tool Claude central to U.S. campaign in Iran, amid a bitter feud","Washington Post","news","","","https://www.washingtonpost.com/technology/2026/03/04/anthropic-ai-iran-campaign/"),
("cornell_lii","Cornell Legal Information Institute","n.d.","Wex Legal Definitions (Natural Person; Artificial Person)","Legal Information Institute","reference_work","","","https://www.law.cornell.edu/wex/natural_person"),
("dodson2023","Dodson, C.","2023","Oral testimony on HB 1361 before the North Dakota Legislative Assembly","North Dakota Legislative Assembly","testimony","ND","HB 1361","https://video.ndlegis.gov/en/PowerBrowser/PowerBrowserV2/20230306/-1/29325"),
("douthat2026","Douthat, R.; Alvarez Boyd, S.","2026","Anthropic's Chief on A.I.: We Don't Know if the Models Are Conscious","New York Times","news","","","https://www.nytimes.com/2026/02/12/opinion/artificial-intelligence-anthropic-amodei.html"),
("earthlaw2024","Earth Law Center","2024","Utah Advances Anti-Rights of Nature Bill with Implications for AI","Earth Law Center","advocacy","UT","HB 249","https://www.earthlawcenter.org/blog-entries/2024/2/utah-advances-anti-rights-of-nature-bill-with-implications-for-ai"),
("edmonson2025","Edmonson, D.","2025","RE: HB 469, concerns (testimony to Ohio House Cttee on Technology and Innovation)","Ohio House of Representatives","testimony","OH","HB 469","https://ohiohouse.gov/committees/technology-and-innovation/bills/hb469"),
("evans2025","Evans, N.","2025","What's in Ohio's Proposal Banning AI Personhood","Ohio Capital Journal","news","OH","HB 469","https://ohiocapitaljournal.com/2025/11/17/whats-in-ohios-proposal-banning-ai-personhood/"),
("fli_a","Future of Life Institute","n.d.","About us","FLI","advocacy","","","https://futureoflife.org/about-us/"),
("fli_b","Future of Life Institute","n.d.","Home","FLI","advocacy","","","https://futureoflife.org/"),
("hochul2025","Governor Kathy Hochul","2025","Governor Hochul Signs Legislation to Require AI Frameworks for AI Frontier Models","NY Governor's Office","press_release","NY","","https://www.governor.ny.gov/news/governor-hochul-signs-nation-leading-legislation-require-ai-frameworks-ai-frontier-models"),
("guardian2022","Guardian","2022","Happy the elephant is not a person, says court in key US animal rights case","The Guardian","news","","","https://www.theguardian.com/us-news/2022/jun/14/elephant-person-animal-rights-happy"),
("hardin2026","Hardin, C.","2026","Governor DeSantis Directs Florida State Agencies to Partner with FLI to Shield Families from AI Harm","Future of Life Institute","press_release","FL","","https://futureoflife.org/press-release/desantis-directs-florida-agencies-to-partner-with-fli/"),
("mo_hb1746_test","HB 1746 testimony before the Missouri House Emerging Issues Committee","2026","Committee testimony record","Missouri House of Representatives","testimony","MO","HB 1746","https://documents.house.mo.gov/billtracking/bills261/witnesses/HB1746Testimony2-2.pdf"),
("mo_hb1769_test","HB 1769 testimony before the Missouri House Emerging Issues Committee","2026","Committee testimony record","Missouri House of Representatives","testimony","MO","HB 1769","https://documents.house.mo.gov/billtracking/bills261/witnesses/HB1769Testimony2-2.pdf"),
("hill2022","Hill, M.","2022","Court must decide if an elephant is a person","Associated Press via KSL","news","","","https://www.ksl.com/article/50407168/court-must-decide-if-an-elephant-is-a-person"),
("iff2022","Idaho Freedom Foundation","2022","House Bill 720 - Personhood, animals, objects","Idaho Freedom Foundation","advocacy","ID","HB 720","https://idahofreedom.org/house-bill-720-personhood-animals-objects/"),
("littleton_nd","Littleton, M.","n.d.","Tennessee House of Representatives member page","Tennessee General Assembly","member_page","TN","HB 1455","https://wapp.capitol.tn.gov/apps/LegislatorInfo/member?district=h78"),
("maisonnave2023","Maisonnave et al.","2023","Indigenous leader inspires an Amazon city to grant personhood to an endangered river","AP News","news","","","https://apnews.com/article/brazil-amazon-wari-indigenous-nature-rights-deforestation-68af65663fb7bd1b9d2051ce10c17a46"),
("massey_nd","Massey, B.","n.d.","Tennessee Senate member page","Tennessee General Assembly","member_page","TN","SB 1493","https://www.capitol.tn.gov/senate/archives/107GA/members/s6.html"),
("mnhrd_nd","Minnesota House Research Department","n.d.","State Constitutional Amendments","Minnesota House Research","gov_reference","MN","SF 4114","https://www.house.mn.gov/hrd/pubs/ss/ssconamend.pdf"),
("mbca","Model Business Corporation Act","n.d.","Section 3.02","-","legal_doc","","","https://www.lexisnexis.com/documents/pdf/20080618091347_large.pdf"),
("nichols2025a","Nichols, T.","2025","AI, States' Rights, and My Visit to WallBuilders","Nichols for Idaho (Substack)","blog","ID","HB 720","https://nicholsforidaho.substack.com/p/ai-states-rights-and-my-visit-to"),
("nichols2025b","Nichols, T.","2025","AI, States' Rights, and the Most Basic Property Right: Owning Yourself","Nichols for Idaho (Substack)","blog","ID","HB 720","https://nicholsforidaho.substack.com/p/ai-states-rights-and-the-most-basic"),
("nhrp2025","Nonhuman Rights Project","2025","NhRP argues Utah court must allow HB 249 lawsuit to proceed","Nonhuman Rights Project","advocacy","UT","HB 249","https://www.nonhumanrights.org/blog/hb249-lawsuit-nhrp/"),
("ndcan2023","North Dakota Can","2023","HB 1361 - Definition of Personhood","North Dakota Can","advocacy","ND","HB 1361","https://ndcan.org/house-bill-1361"),
("ndcan_nd","North Dakota Can","n.d.","About","North Dakota Can","advocacy","ND","","https://ndcan.org/about"),
("ndfa_nd","North Dakota Family Alliance","n.d.","North Dakota Family Alliance","NDFA","advocacy","ND","HB 1361","https://www.ndfamilyalliance.org/"),
("ndleg2023","North Dakota Legislative Assembly","2023","Testimony on and hearing materials on HB 1361","North Dakota Legislative Assembly","testimony","ND","HB 1361","https://ndlegis.gov/sites/default/files/resource/68-2023/library/hb1361.pdf"),
("ndleg2025","North Dakota Legislative Assembly","2025","North Dakota HB 1145 Bill Text","North Dakota Legislative Assembly","legal_doc","ND","HB 1145","https://ndlegis.gov/sites/default/files/resource/69-2025/library/hb1145.pdf"),
("ohiohouse_nd","Ohio House of Representatives","n.d.","HB 469, 135th General Assembly","Ohio House Technology and Innovation Cttee","gov_reference","OH","HB 469","https://ohiohouse.gov/committees/technology-and-innovation/bills/hb469"),
("perrigo2025","Perrigo, B.; Chow, A.","2025","Senators Reject 10-Year Ban on State-Level AI Regulation","Time","news","","","https://time.com/7299044/senators-reject-10-year-ban-on-state-level-ai-regulation-in-blow-to-big-tech/"),
("peta2024","PETA","2024","'Tilikum v. SeaWorld' Case Summary","PETA","advocacy","","","https://www.peta.org/features/peta-foundation-legal/case-summaries/tilikum-v-seaworld/"),
("pwh_nd","Protect What's Human","n.d.","Protect What's Human","Organisation site","advocacy","","","https://protectwhatshuman.org/"),
("roche2024","Roche, L. R.","2024","No 'legal personhood' for the Great Salt Lake, Utah legislative committee decides","Deseret News","news","UT","HB 249","https://www.deseret.com/utah/2024/1/23/24048098/great-salt-lake-legal-personhood-utah-legislature/"),
("romero2024","Romero, M.","2024","Denying Great Salt Lake 'personhood' closes an option for saving the lake, Utahns say","Utah News Dispatch","news","UT","HB 249","https://utahnewsdispatch.com/2024/01/23/great-salt-lake-personhood-bill/"),
("roy2017","Roy, E. A.","2017","New Zealand river granted same legal rights as human being","The Guardian","news","","","https://www.theguardian.com/world/2017/mar/16/new-zealand-river-granted-same-legal-rights-as-human-being"),
("sogsl_nd","Save Our Great Salt Lake","n.d.","HOME","Organisation site","advocacy","UT","HB 249","https://www.saveourgreatsaltlake.org/"),
("tn_sa0922","SB 837 amendment no. 1 (SA0922)","2026","114th Gen. Assemb., 1st Reg. Sess. (Tenn.)","Tennessee General Assembly","legal_doc","TN","HB 849 / SB 837","https://www.capitol.tn.gov/Bills/114/Amend/SA0922.pdf"),
("scott2025","Scott, C.","2025","Testimony on Ohio HB 469","Ohio House of Representatives","testimony","OH","HB 469","https://ohiohouse.gov/committees/technology-and-innovation/bills/hb469"),
("saet_nd","Society for Ancient and Evangelical Theology","n.d.","Image of God (Imago Dei)","SAET","primary_religious","","","https://www.saet.ac.uk/Christianity/ImageofGod"),
("steinhauser_bio","Steinhauser, B.","n.d.","Brendan Steinhauser - Public Affairs and Communications Strategist","LinkedIn","bio","","","https://www.linkedin.com/in/brendanstteinhauser/details/experience/"),
("steinhauser2025","Steinhauser, B.","2025","Protecting Human Agency: HB 469 Testimony","Alliance for Secure AI Action","testimony","OH","HB 469","https://secureainow.org/ohiohb469/"),
("tan2025","Tan, G.","2025","Tweet: TN SB 1493 is a case study in the idiocy that will destroy AI innovation in America","X (Twitter)","social","TN","HB 1455 / SB 1493","https://x.com/garrytan/status/2005002574697824594"),
("technet_nd","TechNet","n.d.","TechNet - The Voice of American Innovation","TechNet","advocacy","","","https://www.technet.org/"),
("tn_ha1260","Tennessee Amendment No. 4 to HB 1455 / SB 1493 (HA1260)","2026","114th Gen. Assemb., 1st Reg. Sess.","Tennessee General Assembly","legal_doc","TN","HB 1455 / SB 1493","https://capitol.tn.gov/Bills/114/Amend/HA1260.pdf"),
("tnga_nd","Tennessee General Assembly","n.d.","Tennessee Code Annotated Section 40-35-111","Tennessee General Assembly","legal_doc","TN","HB 1455","https://law.justia.com/codes/tennessee/title-40/chapter-35/part-1/section-40-35-111/"),
("prohuman2026","The Pro-Human AI Declaration","2026","The Pro-Human AI Declaration","humanstatement.org","advocacy","","","https://humanstatement.org"),
("whitehouse2025","The White House","2025","White House Unveils America's AI Action Plan","The White House","press_release","","","https://www.whitehouse.gov/articles/2025/07/white-house-unveils-americas-ai-action-plan/"),
("ut_house2024","Utah House Hearing","2024","Utah HB 249 House Hearing","Utah State Legislature","testimony","UT","HB 249","https://le.utah.gov/av/committeeArchive.jsp?timelineID=241561"),
("ut_senate2024","Utah Senate Hearing","2024","Utah Senate Judiciary Committee - HB 249 Hearing","Utah State Legislature","testimony","UT","HB 249","https://le.utah.gov/av/committeeArchive.jsp?timelineID=246931"),
("wallbuilders2024","WallBuilders","2024","WallBuilders 2024 Impact Report","WallBuilders","advocacy","","","https://wallbuilders.com/wp-content/uploads/2024/12/2024-WallBuilders-Impact-Report.pdf"),
("wallbuilders2025","WallBuilders","2025","WallBuilders 2025 Impact Report","WallBuilders","advocacy","","","https://wallbuilders.com/wp-content/uploads/2025/12/2025-WallBuilders-Impact-Report.pdf"),
("wallbuilders_nd","WallBuilders","n.d.","Preserve America's Heritage: Education, Resources and Events","WallBuilders","advocacy","","","https://wallbuilders.com/"),
("weller2017","Weller, C.","2017","A robot has just been granted citizenship of Saudi Arabia","World Economic Forum","news","","","https://www.weforum.org/stories/2017/10/a-robot-has-just-been-granted-citizenship-of-saudi-arabia/"),
("weyand2025","Weyand, Z.","2025","Ohio House Bill Introduced to Prohibit Personhood of AI","The Post (Ohio University)","news","OH","HB 469","https://www.thepostathens.com/article/2025/11/ohio-house-bill-469-prohibit-ai-personhood"),
("williams2024","Williams, C.","2024","What is a person? Here's why Utah is considering a new bill to define it","KSL News","news","UT","HB 249","https://www.ksl.com/article/50853851/what-is-a-person-heres-why-utah-is-considering-a-new-bill-to-define-it"),
("wiseye2026","WisconsinEye","2026","Assembly Committee on Science, Technology, and AI","WisconsinEye","testimony","WI","AB 959 / SB 932","https://wiseye.org/2026/02/04/assembly-committee-on-science-technology-and-ai-7/"),
("wef2020","World Economic Forum","2020","What if nature became a legal person?","World Economic Forum","advocacy","","","https://www.weforum.org/stories/2020/05/nature-legal-personhood/"),
("yurok2019","Yurok Tribe","2019","Resolution Establishing the Rights of the Klamath River (Resolution No. 19-40)","Yurok Tribe","legal_doc","","","https://ecojurisprudence.org/initiatives/resolution-establishing-rights-of-the-klamath-river/"),
]

CASES=[
("Ban v. Utah","No. 250900869","Utah Dist. Ct.","2025","NhRP challenge to HB 249; dismissed 8 May 2025"),
("Bank Markazi v. Peterson","578 U.S. 212","U.S. Supreme Court","2016","Congress may amend law affecting pending litigation"),
("Citizens United v. FEC","558 U.S. 310","U.S. Supreme Court","2010","First Amendment protection for corporate political spending"),
("Cleveland Board of Education v. LaFleur","414 U.S. 632","U.S. Supreme Court","1974","Irrebuttable presumptions and due process"),
("Coates v. Cincinnati","402 U.S. 611","U.S. Supreme Court","1971","Vagueness: criminalising 'annoying' behaviour"),
("Connally v. General Construction Co.","269 U.S. 385","U.S. Supreme Court","1926","Vagueness: 'current rate of per diem wages'"),
("Daubert v. Merrell Dow Pharmaceuticals","509 U.S. 579","U.S. Supreme Court","1993","Admissibility standard for scientific evidence"),
("Epperson v. Arkansas","393 U.S. 97","U.S. Supreme Court","1968","Establishment Clause; teaching of evolution"),
("FCC v. Beach Communications","508 U.S. 307","U.S. Supreme Court","1993","Rational basis test"),
("Kennedy v. Bremerton School District","597 U.S. 507","U.S. Supreme Court","2022","Free Exercise; personal religious observance"),
("Kitzmiller v. Dover Area School District","400 F. Supp. 2d 707","M.D. Pa.","2005","Establishment Clause; intelligent design"),
("Marbury v. Madison","5 U.S. (1 Cranch) 137","U.S. Supreme Court","1803","Judicial review; right to claim protection of the laws"),
("Papachristou v. City of Jacksonville","405 U.S. 156","U.S. Supreme Court","1972","Void for vagueness; vagrancy laws"),
("United States v. Carolene Products Co.","304 U.S. 144","U.S. Supreme Court","1938","Presumption of facts supporting legislative judgment"),
("United States v. Klein","80 U.S. (13 Wall.) 128","U.S. Supreme Court","1871","Congress may not prescribe a rule of decision"),
]

# ---------------- styling ----------------
ARIAL=lambda **k: Font(name="Arial", **k)
HDR_FILL=PatternFill("solid", fgColor="1F3864")
BAND=PatternFill("solid", fgColor="F2F5FA")
THIN=Side(style="thin", color="D9D9D9")
BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

def write_sheet(ws, headers, rows, widths, wrap_cols=(), link_cols=()):
    for j,h in enumerate(headers,1):
        c=ws.cell(1,j,h); c.font=ARIAL(bold=True,color="FFFFFF",size=10)
        c.fill=HDR_FILL; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BORDER
    for i,row in enumerate(rows,2):
        for j,v in enumerate(row,1):
            c=ws.cell(i,j,v); c.font=ARIAL(size=10); c.border=BORDER
            c.alignment=Alignment(vertical="top", wrap_text=(headers[j-1] in wrap_cols))
            if i%2==0: c.fill=BAND
            if headers[j-1] in link_cols and isinstance(v,str) and v.startswith("http"):
                c.hyperlink=v; c.font=ARIAL(size=10, color="0563C1", underline="single")
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[1].height=32
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

wb=Workbook(); wb.remove(wb.active)
wb.calculation.fullCalcOnLoad=True  # no LibreOffice here; force recalc when opened

# ---- README ----
ws=wb.create_sheet("README")
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=104
ws["A1"]="AI Legal Status Tracker - Reference Register"; ws["A1"].font=ARIAL(bold=True,size=14,color="1F3864")
ws["A2"]="Phase 1, steps 1-2: reference extraction and DM corpus dedupe"; ws["A2"].font=ARIAL(size=10,italic=True)
readme=[
("Source document","Smith, A., Caviola, L., & Alexander, H. (2026). Denying Personhood to AI: An Analysis of U.S. State Legislation on AI Legal Status. SSRN 6829981. Reference list pp. 42-50, Table of Cases p. 41."),
("Dedupe target","digital-minds-corpus/master-corpus.csv - layers core + canon (1,279 works). Matched on normalised title, threshold 0.85, containment-boosted. Zero ambiguous matches."),
("",""),
("SHEET: Literature","The working set for the literature review. Combines (a) the scholarly subset of the paper's references and (b) the personhood / legal-status cluster already held in the DM corpus but not cited by the paper. Filter on 'in_paper' and 'in_corpus' to separate the two."),
("SHEET: Evidence","Non-scholarly reference entries - testimony, news, advocacy, legislature records, social posts. These are evidence ABOUT the bills, not literature. They are source records for the bill registry in Phase 2, keyed by jurisdiction and related_bill."),
("SHEET: Cases","Table of Cases from the paper (p. 41). Primary law, not literature; relevant to the legal-vulnerabilities analysis."),
("SHEET: Summary","Counts, computed by formula from the sheets above."),
("",""),
("Column: tier","core = substantive scholarship for the review. marginal = think-tank reports, trade commentary and industry essays; context rather than literature. Decide at the checkpoint whether marginal enters the registry."),
("Column: strand","Descriptive topic tag. Flat, non-hierarchical, no ranking - consistent with the Observatory descriptive-first principle."),
("Column: action","'No action - already held' = in the corpus. 'Collect' = new, core tier, obtain in step 3. 'Hold (marginal)' = park pending checkpoint. 'Classify for tracker' = corpus-held work the paper missed."),
("Column: cited_by","Citation count as recorded in the DM corpus at its own as-of date. Blank for works not in the corpus. Not independently re-verified here."),
("",""),
("Editable cells","The 'notes' column on Literature and Evidence is intended for reviewer comment. Everything else is generated - re-run build_workbook.py rather than editing in place."),
("Assumptions","(1) Op-eds by parties to the debate (Berg & Rosenblatt 2025; Alexander & Simon 2025) are classed as evidence, not literature, because they argue about the bills rather than contribute peer-reviewed analysis. (2) SEP entries and the International AI Safety Report are treated as scholarship. (3) Aquinas and Genesis are primary religious sources, not literature."),
("Google Sheets","This workbook converts cleanly to a native Google Sheet: upload to Drive and open with Sheets, or File > Import. All five tabs, the filters, the frozen headers and the Summary formulas survive conversion; Sheets recalculates the Summary on import. Per-sheet CSVs are also provided in csv/ as a fallback - import those as separate tabs named Literature, Evidence and Cases if you want the Summary formulas to resolve."),
("Provenance","Reference data transcribed from the source PDF by hand; corpus fields read from master-corpus.csv. Titles and URLs not re-verified against publishers - do that during step 3 collection."),
]
r=4
for k,v in readme:
    ws.cell(r,1,k).font=ARIAL(bold=bool(k),size=10)
    c=ws.cell(r,2,v); c.font=ARIAL(size=10); c.alignment=Alignment(wrap_text=True,vertical="top")
    r+=1

# ---- Literature ----
LH=["ref_key","in_paper","in_corpus","corpus_layer","tier","strand","authors","year","title","venue","type","discipline","themes","cited_by","url_or_id","action","notes"]
lit.sort(key=lambda x:(x["in_paper"]=="No", x["title"].lower()))
ws=wb.create_sheet("Literature")
write_sheet(ws, LH, [[x[k] for k in LH] for x in lit],
    [14,9,10,12,9,26,30,7,58,32,13,22,30,9,44,24,26], wrap_cols=("title","themes","strand","authors","venue","url_or_id","action","notes"), link_cols=("url_or_id",))

# ---- Evidence ----
EH=["ref_key","authors_or_source","year","title","outlet","evidence_type","jurisdiction","related_bill","url"]
ws=wb.create_sheet("Evidence")
write_sheet(ws, EH, [list(e) for e in sorted(EV, key=lambda x:x[1].lower())],
    [22,32,8,62,30,16,12,20,50], wrap_cols=("title","outlet","authors_or_source","url"), link_cols=("url",))

# ---- Cases ----
CH=["case","citation","court","year","relevance_in_paper"]
ws=wb.create_sheet("Cases")
write_sheet(ws, CH, [list(c) for c in CASES], [40,26,22,8,54], wrap_cols=("relevance_in_paper","case"))

# ---- Summary (formulas) ----
ws=wb.create_sheet("Summary")
ws.column_dimensions["A"].width=52; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=60
ws["A1"]="Summary"; ws["A1"].font=ARIAL(bold=True,size=14,color="1F3864")
NL=len(lit)+1; NE=len(EV)+1; NC=len(CASES)+1
rows=[
("Literature working set (total rows)",f"=COUNTA(Literature!A2:A{NL})","Union of paper scholarly refs and corpus personhood cluster"),
("  Cited by the paper",f'=COUNTIF(Literature!B2:B{NL},"Yes")',"in_paper = Yes"),
("  Already held in DM corpus",f'=COUNTIF(Literature!C2:C{NL},"Yes")',"in_corpus = Yes"),
("  Needs verification",f'=COUNTIF(Literature!C2:C{NL},"Verify")',"Possible retitled duplicate"),
("  Paper refs already in corpus",f'=COUNTIFS(Literature!B2:B{NL},"Yes",Literature!C2:C{NL},"Yes")',"The dedupe saving"),
("  Paper refs NOT in corpus (to collect)",f'=COUNTIFS(Literature!B2:B{NL},"Yes",Literature!C2:C{NL},"No")',"New acquisitions"),
("  Corpus works the paper missed",f'=COUNTIFS(Literature!B2:B{NL},"No",Literature!C2:C{NL},"Yes")',"Finding A: corpus as a source"),
("",None,""),
("Tier: core",f'=COUNTIF(Literature!E2:E{NL},"core")',""),
("Tier: marginal",f'=COUNTIF(Literature!E2:E{NL},"marginal")',""),
("",None,""),
("Action: no action - already held",f'=COUNTIF(Literature!P2:P{NL},"No action - already held")',""),
("Action: verify against corpus record",f'=COUNTIF(Literature!P2:P{NL},"Verify against corpus record")',""),
("Action: collect",f'=COUNTIF(Literature!P2:P{NL},"Collect")',""),
("Action: hold (marginal)",f'=COUNTIF(Literature!P2:P{NL},"Hold (marginal)")',""),
("Action: classify for tracker",f'=COUNTIF(Literature!P2:P{NL},"Classify for tracker")',""),
("",None,""),
("Evidence records (total)",f"=COUNTA(Evidence!A2:A{NE})","Source records for the Phase 2 bill registry"),
("  Testimony",f'=COUNTIF(Evidence!F2:F{NE},"testimony")',""),
("  News",f'=COUNTIF(Evidence!F2:F{NE},"news")',""),
("  Advocacy",f'=COUNTIF(Evidence!F2:F{NE},"advocacy")',""),
("  Legal documents",f'=COUNTIF(Evidence!F2:F{NE},"legal_doc")',""),
("  Op-eds",f'=COUNTIF(Evidence!F2:F{NE},"op_ed")',""),
("  Social posts",f'=COUNTIF(Evidence!F2:F{NE},"social")',""),
("  Press releases",f'=COUNTIF(Evidence!F2:F{NE},"press_release")',""),
("  Other (media, bio, reference, religious, gov)",f'=COUNTA(Evidence!A2:A{NE})-SUM(B19:B25)',""),
("",None,""),
("Cases in Table of Cases",f"=COUNTA(Cases!A2:A{NC})",""),
("",None,""),
("Total reference-list entries accounted for",f'=COUNTIFS(Literature!B2:B{NL},"Yes")+COUNTA(Evidence!A2:A{NE})',"Scholarly + evidence = the paper's reference list"),
("Scholarly share of reference list",f'=COUNTIFS(Literature!B2:B{NL},"Yes")/(COUNTIFS(Literature!B2:B{NL},"Yes")+COUNTA(Evidence!A2:A{NE}))',"Scholarly / (scholarly + evidence)"),
]
r=3
for label,f,note in rows:
    ws.cell(r,1,label).font=ARIAL(size=10,bold=label.startswith(("Literature","Evidence","Total","Cases","Tier","Action","Scholarly")))
    if f is not None:
        c=ws.cell(r,2,f); c.font=ARIAL(size=10); c.alignment=Alignment(horizontal="center")
        if "share" in label: c.number_format="0.0%"
    c=ws.cell(r,3,note); c.font=ARIAL(size=9,italic=True,color="666666"); c.alignment=Alignment(wrap_text=True,vertical="top")
    r+=1

wb.save("reference-register.xlsx")

# ---- per-sheet CSVs (guaranteed-clean Google Sheets import path) ----
import os
os.makedirs("csv", exist_ok=True)
def dump(name, headers, rows):
    with open(f"csv/{name}.csv","w",newline="") as f:
        w=csv.writer(f); w.writerow(headers); w.writerows(rows)
dump("Literature", LH, [[x[k] for k in LH] for x in lit])
dump("Evidence", EH, [list(e) for e in sorted(EV, key=lambda x:x[1].lower())])
dump("Cases", CH, [list(c) for c in CASES])
print("csv/ written:", sorted(os.listdir("csv")))
print(f"literature rows: {len(lit)}  (paper {len(refs)}, corpus-added {added})")
print(f"evidence rows: {len(EV)}   cases: {len(CASES)}")
print(f"paper reference-list total accounted: {len(refs)-len(NOT_IN_PAPER)+len(EV)}")
